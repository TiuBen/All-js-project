# -*- coding: utf-8 -*-
"""
客运航班检查单 → 串联树形 JSON 转换脚本 v1
==========================================
客运与货运结构不同：
- 三个阶段：一、航空器始发 / 二、航空器过站 / 三、航后阶段
- 每节点含：节点名称 | 类型(客运) | 责任单位 | 标准要求
- 新元素：责任单位(responsible)、类型(category)、座级细分时间要求
- 视频监管项同行挂载（挂到该节点的虚拟辅助下）
"""
import openpyxl
import json
import os
import re
import uuid as uuidlib
from datetime import datetime

ROOT = r"C:\Users\HJW-AMD-PRP\Documents\GitHub\All-js-project"
EXCEL_DIR = os.path.join(ROOT, "packages", "frontend", "flight-dispatch-system", "资料")
OUT_DIR = os.path.join(ROOT, "packages", "backend", "full-web-backend", "V5-dispatch", "data", "checklists")
os.makedirs(OUT_DIR, exist_ok=True)


def clean(s):
    if s is None:
        return ""
    return str(s).replace("\n", " ").strip()


def uid():
    return str(uuidlib.uuid4())


def parse_int(s, default=0):
    m = re.search(r"\d+", s or "")
    return int(m.group()) if m else default


def parse_sheet1_passenger(ws):
    """
    Sheet1 行结构：
      R1: 标题
      R2: 表头（日期/航班号/落地时间/机位/起飞时间(EOBT/CTOT)/检查人）
      R3: 一、航空器始发（阶段标题）
      R5: 节点表头（节点名称|类型|责任单位|标准要求）
      R6+: 节点数据行（A=节点名 B=类型 C=责任单位 D=标准要求 G=视频项）
      R19-20: 续行视频项（无节点名）
      R21: 二、航空器过站
      ...
      R42: 三、航后阶段
    """
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([clean(c) for c in row])

    # 阶段标题正则：一、航空器始发 / 二、航空器过站 / 三、航后阶段
    PHASE_RE = re.compile(r"^(一|二|三)、(.+)$")
    HEADER_RE = re.compile(r"节点名称")

    result = {}            # phase_name -> [main nodes]
    video_group_rows = []  # 分组标题
    phase_order = []

    current_phase = None
    current_node = None
    pending_videos = []    # section 标题后、首个节点前的视频项（挂到该阶段第一个节点）
    current_group = None   # 当前视频监管分组标题

    for idx, row in enumerate(rows, start=1):
        a, b, c, d = row[0], row[1], row[2], row[3] if len(row) > 3 else ""
        g = row[6] if len(row) > 6 else ""

        # 阶段标题
        m = PHASE_RE.match(a)
        if m and "节点" not in a and "节点名称" not in a:
            phase_name = m.group(2).strip()   # 航空器始发 / 航空器过站 / 航后阶段
            if phase_name not in result:
                result[phase_name] = []
                phase_order.append(phase_name)
            current_phase = phase_name
            current_node = None
            pending_videos = []
            continue

        # 视频监管组标题（先于表头判断，因为组标题常与表头同行）
        is_group_title = False
        if "视频监管检查重点" in g:
            video_group_rows.append({"row": idx, "title": g.strip()})
            current_group = g.strip()
            is_group_title = True

        # 节点表头
        if HEADER_RE.search(a):
            continue

        # 无阶段时跳过
        if current_phase is None:
            continue

        node_name = a
        node_type = b
        responsible = c
        std_req = d
        video_desc = g if not is_group_title and len(g) > 3 else ""

        # 新节点行（A 列有节点名）
        if node_name:
            node = {
                "uuid": uid(),
                "source": {"row": idx, "seq": len(result[current_phase]) + 1},
                "type": "main",
                "name": node_name,
                "category": node_type,        # 客运
                "responsible": responsible,    # 责任单位（新元素）
                "desc": std_req,
                "auxiliaries": [],
            }
            result[current_phase].append(node)
            current_node = node

            # 挂载 pending 视频（阶段标题后的视频项）到本阶段第一个节点
            if pending_videos:
                aux = {"uuid": uid(), "type": "auxiliary", "name": "（视频监管）", "desc": "",
                       "source": {"row": idx}, "auxiliary": []}
                for pv in pending_videos:
                    pv_node = {"uuid": uid(), "source": {"row": pv["row"]}, "type": "videoSupervision",
                               "name": "", "desc": pv["desc"]}
                    if pv.get("group"):
                        pv_node["group"] = pv["group"]
                    aux["auxiliary"].append(pv_node)
                node["auxiliaries"].append(aux)
                pending_videos = []
        else:
            # 续行：有视频项则挂到当前节点（无当前节点则进 pending）
            if video_desc:
                if current_node is None:
                    pending_videos.append({"row": idx, "desc": video_desc, "group": current_group})
                    continue
                if not current_node["auxiliaries"] or current_node["auxiliaries"][-1]["name"] != "（视频监管）":
                    current_node["auxiliaries"].append(
                        {"uuid": uid(), "type": "auxiliary", "name": "（视频监管）", "desc": "",
                         "source": {"row": idx}, "auxiliary": []})
                current_node["auxiliaries"][-1]["auxiliary"].append({
                    "uuid": uid(), "source": {"row": idx}, "type": "videoSupervision",
                    "name": "", "desc": video_desc, "group": current_group,
                })

        # 同行视频项挂载（节点行自带视频项）
        if node_name and video_desc:
            if not current_node["auxiliaries"] or current_node["auxiliaries"][-1]["name"] != "（视频监管）":
                current_node["auxiliaries"].append(
                    {"uuid": uid(), "type": "auxiliary", "name": "（视频监管）", "desc": "",
                     "source": {"row": idx}, "auxiliary": []})
            # 记录 group 到最后一个视频项
            vnode = {"uuid": uid(), "source": {"row": idx}, "type": "videoSupervision",
                     "name": "", "desc": video_desc, "group": current_group}
            current_node["auxiliaries"][-1]["auxiliary"].append(vnode)

    return result, phase_order, video_group_rows


def parse_sheet3_notes(ws):
    """Sheet3 是修正版说明文本（含细分类型/责任单位表），按文本保存"""
    lines = []
    for row in ws.iter_rows(values_only=True):
        v = row[0]
        if v is not None:
            lines.append(str(v))
    return "\n".join(lines)


def build_passenger_tree(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {
        "uuid": uid(),
        "category": "客运航班",
        "source": os.path.basename(xlsx_path),
        "generatedAt": datetime.now().isoformat(),
    }
    # Sheet1 是主表（始发/过站/航后 三阶段）
    if "Sheet1" in wb.sheetnames:
        ft, order, vg = parse_sheet1_passenger(wb["Sheet1"])
        # 按阶段顺序输出
        ordered = {}
        for p in order:
            ordered[p] = ft[p]
        data["flightTypes"] = ordered
        data["phaseOrder"] = order
        data["videoSupervisionGroups"] = vg
    if "Sheet3" in wb.sheetnames:
        data["phaseNotes"] = parse_sheet3_notes(wb["Sheet3"])
    return data


def main():
    passenger_path = os.path.join(EXCEL_DIR, "客运航班节点保障及合规性监控检查单.xlsx")
    data = build_passenger_tree(passenger_path)
    out = os.path.join(OUT_DIR, "passenger-checklist.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"→ {out}")
    print(f"uuid: {data['uuid']}")
    print(f"阶段: {data.get('phaseOrder')}")
    for p, nodes in data["flightTypes"].items():
        video_count = sum(len(a["auxiliary"]) for n in nodes for a in n["auxiliaries"])
        responsible_set = set(n.get("responsible", "") for n in nodes)
        print(f"[{p}] 节点 {len(nodes)} | 视频项 {video_count} | 责任单位: {', '.join(x for x in responsible_set if x)}")

    # 打印第一个节点
    first_phase = data["phaseOrder"][0] if data.get("phaseOrder") else None
    if first_phase and data["flightTypes"].get(first_phase):
        print(f"\n=== {first_phase} 节点1 示例 ===")
        print(json.dumps(data["flightTypes"][first_phase][0], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
