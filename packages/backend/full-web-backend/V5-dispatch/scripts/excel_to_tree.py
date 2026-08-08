# -*- coding: utf-8 -*-
"""
Excel 检查单 → 串联树形 JSON 转换脚本 v3
==========================================
按 "1 对应 ABCD，A 对应 abc" 的串联父子关系生成：

  main (主监控指标)
   └─ auxiliaries[]: auxiliary (辅助监控指标)
       └─ auxiliary[]: videoSupervision (视频监管重点，挂在与它同行的辅助项下)

每个节点带 uuid（前端填写状态绑定用）与 source（来源行号）。
"""
import openpyxl
import json
import os
import re
import uuid as uuidlib
from datetime import datetime

ROOT = r"C:\Users\HJW-AMD-PRP\Documents\GitHub\All-js-project"
EXCEL_DIR = os.path.join(ROOT, "packages", "frontend", "flight-dispatch-system", "资料")
OUT_DIR = os.path.join(ROOT, "packages", "backend", "full-web-backend", "V5-dispatch", "data", "checklists")
os.makedirs(OUT_DIR, exist_ok=True)


def clean(s):
    if s is None:
        return ""
    return str(s).replace("\n", " ").strip()


def uid():
    return str(uuidlib.uuid4())


def parse_int(s, default=0):
    m = re.search(r"\d+", s or "")
    return int(m.group()) if m else default


# ============================================================
# Sheet1 解析：生成 常规航班 / 始发航班 的串联树
# ============================================================
def parse_sheet1_tree(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([clean(c) for c in row])

    result = {"常规航班": [], "始发航班": []}
    video_group_rows = []   # [(row, title)]

    section = None
    current_main = None       # 当前 main 节点
    current_aux = None        # 当前 auxiliary 节点（用于挂视频项）
    current_video_group = None

    def new_aux(name, desc, row_idx, main_node):
        """在 main_node.auxiliaries 里新增一个 auxiliary 节点"""
        aux = {
            "uuid": uid(),
            "type": "auxiliary",
            "name": name,
            "desc": desc,
            "source": {"row": row_idx},
            "auxiliary": [],  # 挂视频监管重点
        }
        main_node["auxiliaries"].append(aux)
        return aux

    for idx, row in enumerate(rows, start=1):
        a, b, c = row[0], row[1], row[2] if len(row) > 2 else ""
        e, f = (row[4] if len(row) > 4 else ""), (row[5] if len(row) > 5 else "")
        i = row[8] if len(row) > 8 else ""

        # 段落切换
        if a == "始发航班" and idx > 10:
            section = "始发航班"
            current_main = None
            current_aux = None
            continue
        if a == "序号" and "主要监控指标" in b:
            section = section or "常规航班"
            current_main = None
            current_aux = None
            continue

        # 视频监管分组标题（"视频监管检查重点（xxx）"）
        if ("视频监管" in i or "视频监管" in b) and ("（" in i or "（" in b or "检查" in i):
            title = i if "视频监管" in i else b
            current_video_group = title.strip()
            video_group_rows.append({"row": idx, "title": title.strip()})
            # 该行如果同时也是 main/aux 数据行，继续处理（不 return）
        elif "视频监管" in i:
            current_video_group = i.strip()
            video_group_rows.append({"row": idx, "title": i.strip()})

        if section is None:
            continue

        seq = a
        main_name = b
        main_desc = c
        aux_name = e
        aux_desc = f
        video_desc = i

        # 视频项是否为"组标题"（需要跳过，不当作项）
        is_group_title = ("视频监管检查重点" in video_desc) and (video_desc == current_video_group)

        # 1) 新主节点：A 列有序号 且 B 列有名称
        row_has_aux = bool(aux_name)
        if seq and main_name:
            main_node = {
                "uuid": uid(),
                "source": {"row": idx, "seq": parse_int(seq, idx)},
                "type": "main",
                "name": main_name,
                "desc": main_desc,
                "auxiliaries": [],
            }
            result[section].append(main_node)
            current_main = main_node
            current_aux = None
            if aux_name:
                current_aux = new_aux(aux_name, aux_desc, idx, current_main)
        else:
            # 2) 续行：新增 auxiliary（挂在当前 main 下）
            if aux_name and current_main is not None:
                current_aux = new_aux(aux_name, aux_desc, idx, current_main)

        # 3) 视频项挂载：视频项与辅助项按"同行对应"原则
        #    - 本行有辅助项 → 挂到本行的辅助项下
        #    - 本行无辅助项 → 创建虚拟辅助节点（归属主节点）承载
        if video_desc and not is_group_title and len(video_desc) > 3:
            video_node = {
                "uuid": uid(),
                "source": {"row": idx},
                "type": "videoSupervision",
                "name": "",
                "desc": video_desc,
            }
            if current_video_group:
                video_node["group"] = current_video_group

            if current_main is None:
                continue
            if not row_has_aux or current_aux is None:
                # 无辅助项 → 创建虚拟 auxiliary 承载（name 用主节点名）
                current_aux = new_aux(main_name or "（视频监管）", "", idx, current_main)
            current_aux["auxiliary"].append(video_node)

    return result, video_group_rows


# ============================================================
# Sheet3 考核指标（复用之前的解析逻辑）
# ============================================================
def parse_assessment_sheet3(ws):
    rows = [[clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
    out = {"航空器保障": [], "机坪保障": [], "货物保障": []}
    cargo_items = {}
    in_apron = False

    for idx, row in enumerate(rows, start=1):
        a = row[0]
        if a == "机坪保障":
            in_apron = True
            continue
        if a == "序号" and "指标名称" in (row[1] if len(row) > 1 else ""):
            continue
        if a == "考核指标":
            continue

        if not in_apron:
            if a and re.match(r"^\d+$", a):
                seq = int(a)
                if seq not in out["航空器保障"]:
                    out["航空器保障"].append({
                        "seq": seq,
                        "name": row[1] if len(row) > 1 else "",
                        "desc": row[2] if len(row) > 2 else "",
                        "responsible": row[5] if len(row) > 5 else "",
                    })
            g = row[6] if len(row) > 6 else ""
            if g and re.match(r"^\d+$", g):
                seq = int(g)
                if seq not in cargo_items:
                    cargo_items[seq] = {
                        "seq": seq,
                        "name": row[7] if len(row) > 7 else "",
                        "desc": row[8] if len(row) > 8 else "",
                        "responsible": row[11] if len(row) > 11 else "",
                        "aircraftTypes": [],
                    }
                cargo_items[seq]["aircraftTypes"].append({
                    "type": row[8] if len(row) > 8 else "",
                    "duration": row[9] if len(row) > 9 else "",
                    "shunfeng": row[10] if len(row) > 10 else "",
                })
            elif cargo_items:
                last_seq = max(cargo_items.keys())
                if row[8]:
                    cargo_items[last_seq]["aircraftTypes"].append({
                        "type": row[8] if len(row) > 8 else "",
                        "duration": row[9] if len(row) > 9 else "",
                        "shunfeng": row[10] if len(row) > 10 else "",
                    })
        else:
            if a and re.match(r"^\d+$", a):
                seq = int(a)
                if not any(it.get("seq") == seq for it in out["机坪保障"]):
                    out["机坪保障"].append({
                        "seq": seq,
                        "name": row[1] if len(row) > 1 else "",
                        "desc": row[2] if len(row) > 2 else "",
                        "responsible": row[5] if len(row) > 5 else "",
                    })

    for seq in sorted(cargo_items.keys()):
        out["货物保障"].append(cargo_items[seq])
    return out


# ============================================================
# Sheet4 货物保障
# ============================================================
def parse_sheet4(ws):
    rows = [[clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
    sections = {"过站航班货物保障": [], "始发航班货物保障": []}
    current = None
    for idx, row in enumerate(rows, start=1):
        first = row[0]
        if "过站航班货物保障" in first:
            current = "过站航班货物保障"
            continue
        if "始发航班货物保障" in first:
            current = "始发航班货物保障"
            continue
        if first == "备注":
            continue
        if "序号" in first and "指标名称" in (row[1] if len(row) > 1 else ""):
            continue
        if current is None or not first:
            continue
        seq_num = parse_int(first, idx)
        name = row[1] if len(row) > 1 else ""
        aircraft_type = row[2] if len(row) > 2 else ""
        duration = row[5] if len(row) > 5 else ""
        shunfeng = row[6] if len(row) > 6 else ""
        responsible = row[7] if len(row) > 7 else ""
        if aircraft_type and re.match(r"^B\d+", aircraft_type):
            if sections[current]:
                sections[current][-1].setdefault("aircraftTypes", []).append({
                    "type": aircraft_type, "duration": duration, "shunfeng": shunfeng
                })
        else:
            sections[current].append({
                "seq": seq_num, "name": name,
                "desc": row[2] if len(row) > 2 else "",
                "responsible": responsible,
            })
    return sections


def build_cargo_tree(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {
        "uuid": uid(),
        "category": "货运航班",
        "source": os.path.basename(xlsx_path),
        "generatedAt": datetime.now().isoformat(),
    }
    if "Sheet1" in wb.sheetnames:
        ft, vg_rows = parse_sheet1_tree(wb["Sheet1"])
        data["flightTypes"] = ft
        data["videoSupervisionGroups"] = vg_rows
    if "Sheet2" in wb.sheetnames:
        data["originatingFillForm"] = [[clean(c) for c in row] for row in wb["Sheet2"].iter_rows(values_only=True)]
    if "Sheet3" in wb.sheetnames:
        data["assessment"] = parse_assessment_sheet3(wb["Sheet3"])
    if "Sheet4" in wb.sheetnames:
        data["cargoSupport"] = parse_sheet4(wb["Sheet4"])
    return data


def main():
    cargo_path = os.path.join(EXCEL_DIR, "货运航班节点保障及合规性监控检查单.xlsx")
    data = build_cargo_tree(cargo_path)
    out = os.path.join(OUT_DIR, "cargo-checklist.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"→ {out}")
    print(f"uuid: {data['uuid']}")
    for t, nodes in data["flightTypes"].items():
        aux_count = sum(len(n["auxiliaries"]) for n in nodes)
        video_count = sum(
            len(a["auxiliary"])
            for n in nodes for a in n["auxiliaries"]
        )
        print(f"[{t}] 主节点 {len(nodes)} | 辅助项 {aux_count} | 视频项 {video_count}")

    # 打印第一个主节点的树结构
    print("\n=== 常规航班 节点1 树形示例 ===")
    print(json.dumps(data["flightTypes"]["常规航班"][0], ensure_ascii=False, indent=2)[:1800])


if __name__ == "__main__":
    main()
