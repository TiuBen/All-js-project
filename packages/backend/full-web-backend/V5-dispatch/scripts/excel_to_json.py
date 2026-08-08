# -*- coding: utf-8 -*-
"""
Excel 检查单 → JSON 转换脚本（重写版 v2）
生成 大类别航班 → 节点 → 节点检查项目 → 项目检查项 的层级结构。
"""
import openpyxl
import json
import os
import re
from datetime import datetime

ROOT = r"C:\Users\HJW-AMD-PRP\Documents\GitHub\All-js-project"
EXCEL_DIR = os.path.join(ROOT, "packages", "frontend", "flight-dispatch-system", "资料")
OUT_DIR = os.path.join(ROOT, "packages", "backend", "full-web-backend", "V5-dispatch", "data", "checklists")
os.makedirs(OUT_DIR, exist_ok=True)


def clean(s):
    if s is None:
        return ""
    return str(s).replace("\n", " ").strip()


def get_rows(ws, min_row=1, max_row=None):
    if max_row is None:
        max_row = ws.max_row
    return [[clean(c) for c in row] for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True)]


def parse_int(s, default=0):
    m = re.search(r"\d+", s or "")
    return int(m.group()) if m else default


# ============= Sheet1 解析 =============
def parse_cargo_sheet1(ws):
    """
    Sheet1 行结构：
      第 1 行：标题
      第 2-3 行：表头（日期/航班号/... / 视频监管检查重点 等）
      第 4 行：列表头 (序号/主要监控指标/标准描述/完成情况/辅助监控指标/标准描述/完成情况/备注)
      第 5-34 行：常规航班
      第 35 行：始发航班 标题
      第 37 行：列表头（带 标准完成时间 列）
      第 38-63 行：始发航班 数据
    """
    rows = get_rows(ws)
    result = {"常规航班": [], "始发航班": []}
    video_groups = []

    section = None        # 当前 section
    current_video = None  # 当前视频监管组

    def push_video_item(desc, row_idx):
        if desc and current_video is not None:
            current_video["items"].append({"row": row_idx, "desc": desc})

    for idx, row in enumerate(rows, start=1):
        a, b = row[0], row[1] if len(row) > 1 else ""
        i = row[8] if len(row) > 8 else ""

        # 切换 section
        if a == "始发航班" and idx > 10:
            section = "始发航班"
            continue
        if a == "序号" and "主要监控指标" in b:
            section = section or "常规航班"
            continue

        # 视频监管组标题（仍然要继续处理本行数据）
        video_title = None
        if "视频监管" in i:
            video_title = i
        elif "视频监管" in b:
            video_title = b
        # 只在标题包含分组关键词时创建组（如"入位前检查"），避免出现"视频监管重点"这种空组
        if video_title and ("（" in video_title or "检查" in video_title):
            current_video = {"title": video_title, "items": []}
            video_groups.append(current_video)
            # 标记下次同样的 desc 不要当作 item
            last_video_title = video_title

        if section is None:
            continue

        seq = a
        main_name = b
        main_desc = row[2] if len(row) > 2 else ""
        aux_name = row[4] if len(row) > 4 else ""
        aux_desc = row[5] if len(row) > 5 else ""

        # 视频监管项（I 列）→ 归属到当前主节点的最后一个 aux 上，或附加到上一个 video group
        video_desc = i
        # 如果当前 desc 等于分组标题本身，跳过（不算 item）
        if video_desc and current_video is not None and video_desc != video_title and len(video_desc) > 5:
            push_video_item(video_desc, idx)

        # 关键：seq 存在且 main_name 也有值 → 新主节点
        # 如果只有 seq 而 main_name 为空 → 视为续行（aux 行）
        if seq and main_name:
            seq_num = parse_int(seq, idx)
            node = {
                "row": idx,
                "seq": seq_num,
                "type": "main",
                "name": main_name,
                "desc": main_desc,
                "auxiliaries": [],
            }
            if aux_name:
                node["auxiliaries"].append({
                    "row": idx, "name": aux_name, "desc": aux_desc
                })
            result[section].append(node)
        else:
            # 续行：附加 aux（也可能 seq 存在但 name 为空）
            if aux_name and result[section]:
                result[section][-1]["auxiliaries"].append({
                    "row": idx, "name": aux_name, "desc": aux_desc
                })

    return result, video_groups


# ============= Sheet3 考核指标 解析 =============
def parse_assessment_sheet3(ws):
    """
    Sheet3 行结构：
      第 1 行：标题
      第 2 行：分类标题 航空器保障(A) | 货物保障(G)
      第 3 行：列表头
      第 4 行起：航空器保障数据(A-F) + 货物保障数据(G-L) 并列
      第 23 行：机坪保障 标题 (A列)
      第 24 行：机坪保障列表头
      第 25 行起：机坪保障数据(A-F)
    """
    rows = get_rows(ws)
    out = {"航空器保障": [], "机坪保障": [], "货物保障": []}

    # 第一阶段：航空器保障 + 货物保障（同时进行，按行号）
    # 第二阶段：机坪保障
    in_apron = False
    cargo_items = {}  # seq -> item

    for idx, row in enumerate(rows, start=1):
        a = row[0]
        if a == "机坪保障":
            in_apron = True
            continue
        if a == "序号" and "指标名称" in (row[1] if len(row) > 1 else ""):
            continue
        if a == "考核指标":
            continue

        if not in_apron:
            # 航空器保障 (A-F)
            if a and re.match(r"^\d+$", a):
                seq = int(a)
                if seq not in out["航空器保障"]:
                    out["航空器保障"].append({
                        "seq": seq,
                        "name": row[1] if len(row) > 1 else "",
                        "desc": row[2] if len(row) > 2 else "",
                        "responsible": row[5] if len(row) > 5 else "",
                    })
            # 货物保障 (G-L)
            g = row[6] if len(row) > 6 else ""
            if g and re.match(r"^\d+$", g):
                seq = int(g)
                if seq not in cargo_items:
                    cargo_items[seq] = {
                        "seq": seq,
                        "name": row[7] if len(row) > 7 else "",
                        "desc": row[8] if len(row) > 8 else "",
                        "responsible": row[11] if len(row) > 11 else "",
                        "aircraftTypes": [],
                    }
                # 当前行也可能是机型行
                cargo_items[seq]["aircraftTypes"].append({
                    "type": row[8] if len(row) > 8 else "",
                    "duration": row[9] if len(row) > 9 else "",
                    "shunfeng": row[10] if len(row) > 10 else "",
                })
            elif cargo_items:
                # 续行：把机型数据附加到上一条 cargo item
                last_seq = max(cargo_items.keys())
                if row[8]:  # I 列有内容说明是机型
                    cargo_items[last_seq]["aircraftTypes"].append({
                        "type": row[8] if len(row) > 8 else "",
                        "duration": row[9] if len(row) > 9 else "",
                        "shunfeng": row[10] if len(row) > 10 else "",
                    })
        else:
            # 机坪保障 (A-F)
            if a and re.match(r"^\d+$", a):
                seq = int(a)
                if not any(item.get("seq") == seq for item in out["机坪保障"]):
                    out["机坪保障"].append({
                        "seq": seq,
                        "name": row[1] if len(row) > 1 else "",
                        "desc": row[2] if len(row) > 2 else "",
                        "responsible": row[5] if len(row) > 5 else "",
                    })

    # 把货物保障按 seq 排序后填入 out
    for seq in sorted(cargo_items.keys()):
        out["货物保障"].append(cargo_items[seq])

    return out


# ============= Sheet4 货物保障 =============
def parse_sheet4(ws):
    rows = get_rows(ws)
    sections = {"过站航班货物保障": [], "始发航班货物保障": []}
    current = None
    for idx, row in enumerate(rows, start=1):
        first = row[0]
        if "过站航班货物保障" in first:
            current = "过站航班货物保障"
            continue
        if "始发航班货物保障" in first:
            current = "始发航班货物保障"
            continue
        if first == "备注":
            continue
        if "序号" in first and "指标名称" in (row[1] if len(row) > 1 else ""):
            continue
        if current is None or not first:
            continue
        seq_num = parse_int(first, idx)
        name = row[1] if len(row) > 1 else ""
        aircraft_type = row[2] if len(row) > 2 else ""
        duration = row[5] if len(row) > 5 else ""
        shunfeng = row[6] if len(row) > 6 else ""
        responsible = row[7] if len(row) > 7 else ""

        if aircraft_type and re.match(r"^B\d+", aircraft_type):
            if sections[current]:
                sections[current][-1].setdefault("aircraftTypes", []).append({
                    "type": aircraft_type, "duration": duration, "shunfeng": shunfeng
                })
        else:
            sections[current].append({
                "seq": seq_num,
                "name": name,
                "desc": row[2] if len(row) > 2 else "",
                "responsible": responsible,
            })
    return sections


def build_checklist_json(category, xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {
        "category": category,
        "source": os.path.basename(xlsx_path),
        "generatedAt": datetime.now().isoformat(),
    }
    if "Sheet1" in wb.sheetnames:
        s1 = wb["Sheet1"]
        ft, vg = parse_cargo_sheet1(s1)
        data["flightTypes"] = ft
        data["videoSupervision"] = vg
    if "Sheet2" in wb.sheetnames:
        data["originatingFillForm"] = get_rows(wb["Sheet2"])
    if "Sheet3" in wb.sheetnames:
        data["assessment"] = parse_assessment_sheet3(wb["Sheet3"])
    if "Sheet4" in wb.sheetnames:
        data["cargoSupport"] = parse_sheet4(wb["Sheet4"])
    return data


def main():
    cargo_path = os.path.join(EXCEL_DIR, "货运航班节点保障及合规性监控检查单.xlsx")
    passenger_path = os.path.join(EXCEL_DIR, "客运航班节点保障及合规性监控检查单.xlsx")

    for path, cat in [(cargo_path, "货运航班"), (passenger_path, "客运航班")]:
        if not os.path.exists(path):
            continue
        print(f"\n=== {cat}: {os.path.basename(path)} ===")
        data = build_checklist_json(cat, path)
        suffix = "cargo" if "货运" in cat else "passenger"
        out = os.path.join(OUT_DIR, f"{suffix}-checklist.json")
        with open(out, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  -> {out}")
        if "flightTypes" in data:
            for t, nodes in data["flightTypes"].items():
                print(f"  {t}: {len(nodes)} 个主节点")
        if "videoSupervision" in data:
            print(f"  视频监管组数: {len(data['videoSupervision'])}")
            for g in data["videoSupervision"]:
                print(f"    - {g['title']}: {len(g['items'])} 项")
        if "assessment" in data:
            for k, v in data["assessment"].items():
                print(f"  考核-{k}: {len(v)} 项")
        if "cargoSupport" in data:
            for k, v in data["cargoSupport"].items():
                print(f"  货物保障-{k}: {len(v)} 项")


if __name__ == "__main__":
    main()